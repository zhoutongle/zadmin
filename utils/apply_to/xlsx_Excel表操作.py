import random
from openpyxl import Workbook

"""
犯得错: 必须所有写完后再关闭
"""
# write
wb = Workbook()
ws = wb.create_sheet('first', 0)

# 获取excel文件的所有sheet以列表返回
sheet_list = wb.sheetnames

# 获取excel中每个sheet
for i in wb:
    print(i)

# 直接获取指定sheet对象
sheet1 = wb['first']

# 指定单元格写入数据
sheet1['a1'] = '姓名'
sheet1['b1'] = '年龄'

# 或者通过cell函数来实现
sheet1.cell(row=2, column=1, value='xxx')
sheet1.cell(row=2, column=2, value='30')

sheet2 = wb.create_sheet('第二个表格', 1)
# 如果行内有内容，从下一行空白行开始写入
# 一次添加1行写入
row = range(1, 10)
sheet2.append(row)
# 一次添加多行
# rows = [
    # ['姓名', '年龄', '性别'],
    # ['sun', '30', 'male'],
    # ['xiu', '30', 'male'],
    # ['wen', '30', 'male'],
# ]
rows = [
    ['姓名', '年龄', '性别']
]
for i in range(100):
    info = []
    name = ''.join([random.choice([chr(i) for i in range(97, 123)]) for i in range(5)])
    age = random.randint(1, 100)
    sex = random.choice(['male', 'female'])
    info.append(name)
    info.append(age)
    info.append(sex)
    rows.append(info)
    
for i in rows:
    sheet2.append(i)

wb.save('manage1.xlsx')